import openpyxl
import pandas as pd

wb = openpyxl.load_workbook("locators.xlsx")

df = pd.read_excel("locators.xlsx")
rows_with_value = df.index[(df['page name']=='login') & (df['locator name']=='email field') & (df['locator type']=='id')].tolist()
print(df.iloc[rows_with_value[0]]["locator"])
# sheet = wb.active
# sheet = wb.worksheets[0]
# first_row = []
# # row = sheet.max_row
# # column = sheet.max_column
# for rows in sheet.iter_cols():
#
#     column_name = rows[0].value
#     print(rows)




